"""
Excel output generation module.
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Handles generation of Excel output files."""

    def __init__(self):
        """Initialize the ExcelGenerator."""
        self.base_letter_map = {1: "A", 2: "B", 3: "C"}

    def generate_output(
        self,
        assignments,
        data,
        mechanic_skills,
        mechanic_inspector_skills,
        inspector_req_columns,
        base_schedule_df,
    ):
        """
        Generate Excel output file.

        Args:
            assignments: List of assignment dictionaries
            data: Dictionary containing processed data
            mechanic_skills: Dictionary of mechanic skills
            mechanic_inspector_skills: Dictionary of mechanic inspector skills
            inspector_req_columns: List of inspector requirement columns
            base_schedule_df: DataFrame with base schedule information

        Returns:
            Workbook: OpenPyXL Workbook object
        """
        logger.info("Generating Excel output")
        bases = data["bases"]
        periods = data["periods"]
        shifts = data["shifts"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Roster"

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = ["Base", "Roster", "Airframe", "Engine", "Avionics", "Position"] + [
            f"Day {i}" for i in range(1, 31)
        ]
        ws.append(headers)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Helper functions
        def get_position(mechanic_id, has_inspector_skill_for_shift, mechanic_skills_dict):
            # Inspector override: if selected as inspector for this shift, always Inspector
            if has_inspector_skill_for_shift:
                return "Inspector"

            has_airframe = False
            has_engine = False
            has_avionics = False

            for aircraft in ["aw139", "h175", "sk92"]:
                if mechanic_skills_dict.get(f"{aircraft}_af", 0) == 1:
                    has_airframe = True
                if mechanic_skills_dict.get(f"{aircraft}_r", 0) == 1:
                    has_engine = True
                if mechanic_skills_dict.get(f"{aircraft}_av", 0) == 1:
                    has_avionics = True

            # Position rules:
            # - If mechanic has avionics skill and any of airframe/engine missing → Avionic
            # - If mechanic has airframe and/or engine and no avionics → Mechanic
            # - If mechanic has all three skills → Mechanic
            if has_avionics and not (has_airframe and has_engine):
                return "Avionic"
            if has_airframe or has_engine or has_avionics:
                return "Mechanic"
            # If no skills flagged, default to Mechanic
            return "Mechanic"

        def has_skill_type(mechanic_id, skill_type, mechanic_skills_dict):
            for aircraft in ["aw139", "h175", "sk92"]:
                if mechanic_skills_dict.get(f"{aircraft}{skill_type}", 0) == 1:
                    return True
            return False

        # Organize assignments
        assignments_by_base_shift = {}
        for assignment in assignments:
            base_id = assignment["base_id"]
            shift = assignment["shift"]
            key = (self.base_letter_map.get(base_id, base_id), shift)

            if key not in assignments_by_base_shift:
                assignments_by_base_shift[key] = []
            assignments_by_base_shift[key].append(
                {
                    "mechanic_id": assignment["mechanic_id"],
                    "group": assignment["group"],
                    "shift": shift,
                }
            )

        sorted_keys = sorted(assignments_by_base_shift.keys(), key=lambda x: (x[0], x[1]))
        current_row = 2

        for (base_letter, shift_num) in sorted_keys:
            shift_name = "Day Shift" if shift_num == 1 else "Night Shift"

            # Section header
            ws.append([shift_name] + [None] * (len(headers) - 1))
            section_row = current_row
            cell = ws.cell(row=section_row, column=1)
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            for col_idx in range(2, len(headers) + 1):
                cell = ws.cell(row=section_row, column=col_idx)
                cell.border = thin_border
            current_row += 1

            base_assignments = assignments_by_base_shift[(base_letter, shift_num)]
            base_assignments.sort(key=lambda x: x["mechanic_id"])

            for assignment_info in base_assignments:
                mechanic_id = assignment_info["mechanic_id"]
                group = assignment_info["group"]
                m_skills = mechanic_skills.get(mechanic_id, {})

                # Determine if this mechanic is acting as inspector
                is_inspector = False
                if inspector_req_columns:
                    for _, row in base_schedule_df.iterrows():
                        if int(row["base_id"]) == [
                            k for k, v in self.base_letter_map.items() if v == base_letter
                        ][0] and int(row["shift"]) == shift_num:
                            for inspector_col in inspector_req_columns:
                                if inspector_col in base_schedule_df.columns:
                                    inspector_required = row[inspector_col]
                                    if pd.notna(inspector_required) and inspector_required > 0:
                                        if mechanic_inspector_skills.get(mechanic_id, {}).get(
                                            inspector_col, 0
                                        ) == 1:
                                            regular_skill_name = inspector_col.replace("_inspec", "")
                                            other_mechanics = [
                                                a["mechanic_id"]
                                                for a in base_assignments
                                                if a["mechanic_id"] != mechanic_id
                                                and mechanic_skills.get(a["mechanic_id"], {}).get(
                                                    regular_skill_name, 0
                                                )
                                                == 1
                                            ]
                                            if other_mechanics:
                                                is_inspector = True
                                                break
                            if is_inspector:
                                break

                position = get_position(mechanic_id, is_inspector, m_skills)
                has_airframe = has_skill_type(mechanic_id, "_af", m_skills)
                has_engine = has_skill_type(mechanic_id, "_r", m_skills)
                has_avionics = has_skill_type(mechanic_id, "_av", m_skills)

                # Row data with checkbox-style symbols
                row_data = [
                    base_letter,
                    f"Mechanic {mechanic_id}",
                    "☑" if has_airframe else "☐",
                    "☑" if has_engine else "☐",
                    "☑" if has_avionics else "☐",
                    position,
                ]

                for day in range(1, 31):
                    if (day <= 15 and group == 1) or (day > 15 and group == 2):
                        row_data.append("☑")
                    else:
                        row_data.append("☐")

                ws.append(row_data)

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.alignment = center_align
                    cell.border = thin_border

                current_row += 1

        for col_idx, header in enumerate(headers, 1):
            if col_idx <= 6:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 8

        logger.info("Excel output generated successfully")
        return wb
