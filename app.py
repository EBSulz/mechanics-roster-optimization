"""
Mechanics Roster Optimization - Streamlit App
This app allows users to upload Excel files and generate optimized roster assignments.
"""

import streamlit as st
import pandas as pd
import numpy as np
from ortools.linear_solver import pywraplp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import time
from datetime import datetime

# Page configuration
st.set_page_config(
    page_title="Mechanics Roster Optimization",
    page_icon="🔧",
    layout="wide",
)

# Custom CSS for better styling
st.markdown(
    """
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def load_data(mechanic_skills_file, base_schedule_file, cost_matrix_file, avoidance_list_file):
    """Load and process all input data files."""
    data = {}

    # Load mechanic skills dataset
    mechanic_skills_df = pd.read_excel(mechanic_skills_file)
    mechanics = sorted(mechanic_skills_df["mechanic_id"].unique().tolist())
    data["mechanic_skills_df"] = mechanic_skills_df
    data["mechanics"] = mechanics

    # Load base aircraft schedule
    base_schedule_df = pd.read_excel(base_schedule_file)
    bases = sorted(base_schedule_df["base_id"].unique().tolist())
    periods = sorted(base_schedule_df["period"].unique().tolist())
    shifts = sorted(base_schedule_df["shift"].unique().tolist())
    data["base_schedule_df"] = base_schedule_df
    data["bases"] = bases
    data["periods"] = periods
    data["shifts"] = shifts

    # Load cost matrix
    cost_matrix_df = pd.read_excel(cost_matrix_file)
    base_column_mapping = {"A": 1, "B": 2, "C": 3}
    cost_dict = {}
    for _, row in cost_matrix_df.iterrows():
        mechanic_id = int(row["id"])
        for col, base_id in base_column_mapping.items():
            if col in cost_matrix_df.columns:
                cost_dict[(mechanic_id, base_id)] = float(row[col])
    data["cost_matrix_df"] = cost_matrix_df
    data["cost_dict"] = cost_dict
    data["base_column_mapping"] = base_column_mapping

    # Load avoidance list (optional)
    avoidance_dict = {}
    if avoidance_list_file is not None:
        try:
            avoidance_df = pd.read_excel(avoidance_list_file)
            for _, row in avoidance_df.iterrows():
                m1 = int(row["mechanic_id"])
                m2 = int(row["avoid_mechanic_id"])
                penalty = float(row["penalty"])
                avoidance_dict[(m1, m2)] = penalty
                avoidance_dict[(m2, m1)] = penalty
        except Exception as e:
            st.warning(f"Could not load avoidance list: {e}")
    data["avoidance_dict"] = avoidance_dict

    return data


def create_model(data):
    """Create and solve the optimization model."""
    # Extract data
    mechanic_skills_df = data["mechanic_skills_df"]
    base_schedule_df = data["base_schedule_df"]
    mechanics = data["mechanics"]
    bases = data["bases"]
    periods = data["periods"]
    shifts = data["shifts"]
    cost_dict = data["cost_dict"]
    avoidance_dict = data["avoidance_dict"]

    # Create solver
    solver = pywraplp.Solver.CreateSolver("SCIP")
    if not solver:
        solver = pywraplp.Solver.CreateSolver("CBC")

    # Create decision variables: x[m, b, g, s]
    x = {}
    for m in mechanics:
        for b in bases:
            for g in periods:
                for s in shifts:
                    var_name = f"x_m{m}_b{b}_g{g}_s{s}"
                    x[(m, b, g, s)] = solver.IntVar(0, 1, var_name)

    # Constraint 1: Each mechanic ≤ 1 assignment
    for m in mechanics:
        constraint = solver.Constraint(0, 1, f"mechanic_{m}_single_assignment")
        for b in bases:
            for g in periods:
                for s in shifts:
                    constraint.SetCoefficient(x[(m, b, g, s)], 1)

    # Constraint 2: Skill coverage
    aircraft_types = ["aw139", "h175", "sk92"]
    skill_types = ["_af", "_r", "_av"]

    mechanic_skills = {}
    mechanic_inspector_skills = {}
    for _, row in mechanic_skills_df.iterrows():
        m = int(row["mechanic_id"])
        mechanic_skills[m] = {}
        mechanic_inspector_skills[m] = {}
        for aircraft in aircraft_types:
            for skill in skill_types:
                col_name = f"{aircraft}{skill}"
                if col_name in mechanic_skills_df.columns:
                    mechanic_skills[m][col_name] = int(row[col_name])
                inspector_col_name = f"{aircraft}{skill}_inspec"
                if inspector_col_name in mechanic_skills_df.columns:
                    mechanic_inspector_skills[m][inspector_col_name] = int(row[inspector_col_name])

    for _, row in base_schedule_df.iterrows():
        base_id = int(row["base_id"])
        period = int(row["period"])
        shift = int(row["shift"])

        for aircraft in aircraft_types:
            if aircraft in base_schedule_df.columns and row[aircraft] > 0:
                for skill in skill_types:
                    skill_name = f"{aircraft}{skill}"
                    constraint = solver.Constraint(
                        1, solver.infinity(), f"skill_{skill_name}_base{base_id}_period{period}_shift{shift}"
                    )
                    for m in mechanics:
                        if mechanic_skills[m].get(skill_name, 0) == 1:
                            constraint.SetCoefficient(x[(m, base_id, period, shift)], 1)

    # Constraint 3: Inspector coverage
    inspector_req_columns = [col for col in base_schedule_df.columns if col.endswith("_inspec")]
    if inspector_req_columns:
        for _, row in base_schedule_df.iterrows():
            base_id = int(row["base_id"])
            period = int(row["period"])
            shift = int(row["shift"])

            for inspector_col in inspector_req_columns:
                if inspector_col in base_schedule_df.columns:
                    inspector_required = row[inspector_col]
                    if pd.notna(inspector_required) and inspector_required > 0:
                        constraint = solver.Constraint(
                            1,
                            solver.infinity(),
                            f"inspector_{inspector_col}_base{base_id}_period{period}_shift{shift}",
                        )
                        for m in mechanics:
                            if mechanic_inspector_skills[m].get(inspector_col, 0) == 1:
                                constraint.SetCoefficient(x[(m, base_id, period, shift)], 1)

    # Constraint 4: No self-inspection
    if inspector_req_columns:
        for _, row in base_schedule_df.iterrows():
            base_id = int(row["base_id"])
            period = int(row["period"])
            shift = int(row["shift"])

            for inspector_col in inspector_req_columns:
                if inspector_col in base_schedule_df.columns:
                    inspector_required = row[inspector_col]
                    if pd.notna(inspector_required) and inspector_required > 0:
                        regular_skill_name = inspector_col.replace("_inspec", "")

                        for m_inspector in mechanics:
                            if mechanic_inspector_skills[m_inspector].get(inspector_col, 0) == 1:
                                other_mechanics_with_skill = [
                                    m
                                    for m in mechanics
                                    if m != m_inspector and mechanic_skills[m].get(regular_skill_name, 0) == 1
                                ]

                                if other_mechanics_with_skill:
                                    constraint = solver.Constraint(
                                        -solver.infinity(),
                                        0,
                                        f"no_self_inspect_{inspector_col}"
                                        f"_base{base_id}_period{period}_shift{shift}_inspector{m_inspector}",
                                    )
                                    constraint.SetCoefficient(x[(m_inspector, base_id, period, shift)], 1)
                                    for m_other in other_mechanics_with_skill:
                                        constraint.SetCoefficient(x[(m_other, base_id, period, shift)], -1)

    # Avoidance penalty variables
    avoidance_vars = {}
    if avoidance_dict:
        unique_pairs = set()
        for (m1, m2) in avoidance_dict.keys():
            if m1 < m2:
                unique_pairs.add((m1, m2))

        for (m1, m2) in unique_pairs:
            for b in bases:
                for g in periods:
                    for s in shifts:
                        var_name = f"y_avoid_m{m1}_m{m2}_b{b}_g{g}_s{s}"
                        avoidance_vars[(m1, m2, b, g, s)] = solver.IntVar(0, 1, var_name)

        # Linearization constraints
        for (m1, m2) in unique_pairs:
            for b in bases:
                for g in periods:
                    for s in shifts:
                        y_var = avoidance_vars[(m1, m2, b, g, s)]

                        c1 = solver.Constraint(-solver.infinity(), 0, f"avoid_y_le_x1_m{m1}_m{m2}_b{b}_g{g}_s{s}")
                        c1.SetCoefficient(y_var, 1)
                        c1.SetCoefficient(x[(m1, b, g, s)], -1)

                        c2 = solver.Constraint(-solver.infinity(), 0, f"avoid_y_le_x2_m{m1}_m{m2}_b{b}_g{g}_s{s}")
                        c2.SetCoefficient(y_var, 1)
                        c2.SetCoefficient(x[(m2, b, g, s)], -1)

                        c3 = solver.Constraint(-1, solver.infinity(), f"avoid_y_ge_sum_m{m1}_m{m2}_b{b}_g{g}_s{s}")
                        c3.SetCoefficient(y_var, -1)
                        c3.SetCoefficient(x[(m1, b, g, s)], 1)
                        c3.SetCoefficient(x[(m2, b, g, s)], 1)

    # Objective function
    objective = solver.Objective()

    # Movement costs
    for m in mechanics:
        for b in bases:
            cost = cost_dict.get((m, b), 0)
            for g in periods:
                for s in shifts:
                    objective.SetCoefficient(x[(m, b, g, s)], cost)

    # Avoidance penalties
    if avoidance_dict and avoidance_vars:
        unique_pairs = set()
        for (m1, m2) in avoidance_dict.keys():
            if m1 < m2:
                unique_pairs.add((m1, m2))

        for (m1, m2) in unique_pairs:
            penalty = avoidance_dict[(m1, m2)]
            for b in bases:
                for g in periods:
                    for s in shifts:
                        if (m1, m2, b, g, s) in avoidance_vars:
                            objective.SetCoefficient(avoidance_vars[(m1, m2, b, g, s)], penalty)

    objective.SetMinimization()

    return solver, x, mechanic_skills, mechanic_inspector_skills, inspector_req_columns, avoidance_vars


def solve_model(solver, progress_bar, status_text):
    """Solve the optimization model with progress updates."""
    status_text.text("🔄 Solving optimization problem...")
    progress_bar.progress(0.8)

    start_time = time.time()
    status = solver.Solve()
    solve_time = time.time() - start_time

    progress_bar.progress(1.0)

    return status, solve_time


def extract_solution(x, mechanics, bases, periods, shifts, cost_dict):
    """Extract solution from the solved model."""
    assignments = []
    total_cost = 0

    for m in mechanics:
        for b in bases:
            for g in periods:
                for s in shifts:
                    if x[(m, b, g, s)].solution_value() > 0.5:
                        cost = cost_dict.get((m, b), 0)
                        total_cost += cost
                        assignments.append(
                            {
                                "mechanic_id": m,
                                "base_id": b,
                                "group": g,
                                "shift": s,
                                "shift_name": "Day" if s == 1 else "Night",
                                "cost": cost,
                            }
                        )

    return assignments, total_cost


def generate_excel_output(
    assignments, data, mechanic_skills, mechanic_inspector_skills, inspector_req_columns, base_schedule_df
):
    """Generate Excel output file."""
    bases = data["bases"]
    periods = data["periods"]
    shifts = data["shifts"]
    base_letter_map = {1: "A", 2: "B", 3: "C"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = ["Base", "Roster", "Airframe", "Engine", "Avionics", "Position"] + [f"Day {i}" for i in range(1, 31)]
    ws.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Helper functions
    def get_position(mechanic_id, has_inspector_skill_for_shift, mechanic_skills_dict):
        # Inspector override: if selected as inspector for this shift, always Inspector
        if has_inspector_skill_for_shift:
            return "Inspector"

        has_airframe = False
        has_engine = False
        has_avionics = False

        for aircraft in ["aw139", "h175", "sk92"]:
            if mechanic_skills_dict.get(f"{aircraft}_af", 0) == 1:
                has_airframe = True
            if mechanic_skills_dict.get(f"{aircraft}_r", 0) == 1:
                has_engine = True
            if mechanic_skills_dict.get(f"{aircraft}_av", 0) == 1:
                has_avionics = True

        # Position rules:
        # - If mechanic has avionics skill and any of airframe/engine missing → Avionic
        # - If mechanic has airframe and/or engine and no avionics → Mechanic
        # - If mechanic has all three skills → Mechanic
        if has_avionics and not (has_airframe and has_engine):
            return "Avionic"
        if has_airframe or has_engine or has_avionics:
            return "Mechanic"
        # If no skills flagged, default to Mechanic
        return "Mechanic"

    def has_skill_type(mechanic_id, skill_type, mechanic_skills_dict):
        for aircraft in ["aw139", "h175", "sk92"]:
            if mechanic_skills_dict.get(f"{aircraft}{skill_type}", 0) == 1:
                return True
        return False

    # Organize assignments
    assignments_by_base_shift = {}
    for assignment in assignments:
        base_id = assignment["base_id"]
        shift = assignment["shift"]
        key = (base_letter_map.get(base_id, base_id), shift)

        if key not in assignments_by_base_shift:
            assignments_by_base_shift[key] = []
        assignments_by_base_shift[key].append(
            {
                "mechanic_id": assignment["mechanic_id"],
                "group": assignment["group"],
                "shift": shift,
            }
        )

    sorted_keys = sorted(assignments_by_base_shift.keys(), key=lambda x: (x[0], x[1]))
    current_row = 2

    for (base_letter, shift_num) in sorted_keys:
        shift_name = "Day Shift" if shift_num == 1 else "Night Shift"

        # Section header
        ws.append([shift_name] + [None] * (len(headers) - 1))
        section_row = current_row
        cell = ws.cell(row=section_row, column=1)
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for col_idx in range(2, len(headers) + 1):
            cell = ws.cell(row=section_row, column=col_idx)
            cell.border = thin_border
        current_row += 1

        base_assignments = assignments_by_base_shift[(base_letter, shift_num)]
        base_assignments.sort(key=lambda x: x["mechanic_id"])

        for assignment_info in base_assignments:
            mechanic_id = assignment_info["mechanic_id"]
            group = assignment_info["group"]
            m_skills = mechanic_skills.get(mechanic_id, {})

            # Determine if this mechanic is acting as inspector
            is_inspector = False
            if inspector_req_columns:
                for _, row in base_schedule_df.iterrows():
                    if int(row["base_id"]) == [k for k, v in base_letter_map.items() if v == base_letter][0] and int(
                        row["shift"]
                    ) == shift_num:
                        for inspector_col in inspector_req_columns:
                            if inspector_col in base_schedule_df.columns:
                                inspector_required = row[inspector_col]
                                if pd.notna(inspector_required) and inspector_required > 0:
                                    if mechanic_inspector_skills.get(mechanic_id, {}).get(inspector_col, 0) == 1:
                                        regular_skill_name = inspector_col.replace("_inspec", "")
                                        other_mechanics = [
                                            a["mechanic_id"]
                                            for a in base_assignments
                                            if a["mechanic_id"] != mechanic_id
                                            and mechanic_skills.get(a["mechanic_id"], {}).get(regular_skill_name, 0)
                                            == 1
                                        ]
                                        if other_mechanics:
                                            is_inspector = True
                                            break
                        if is_inspector:
                            break

            position = get_position(mechanic_id, is_inspector, m_skills)
            has_airframe = has_skill_type(mechanic_id, "_af", m_skills)
            has_engine = has_skill_type(mechanic_id, "_r", m_skills)
            has_avionics = has_skill_type(mechanic_id, "_av", m_skills)

            # Row data with checkbox-style symbols
            row_data = [
                base_letter,
                f"Mechanic {mechanic_id}",
                "☑" if has_airframe else "☐",
                "☑" if has_engine else "☐",
                "☑" if has_avionics else "☐",
                position,
            ]

            for day in range(1, 31):
                if (day <= 15 and group == 1) or (day > 15 and group == 2):
                    row_data.append("☑")
                else:
                    row_data.append("☐")

            ws.append(row_data)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.alignment = center_align
                cell.border = thin_border

            current_row += 1

    for col_idx, header in enumerate(headers, 1):
        if col_idx <= 6:
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 8

    return wb


def main():
    """Main Streamlit app."""
    st.markdown(
        '<h1 class="main-header">🔧 Mechanics Roster Optimization</h1>',
        unsafe_allow_html=True,
    )

    st.sidebar.header("📁 Upload Files")

    mechanic_skills_file = st.sidebar.file_uploader(
        "Mechanic Skills Dataset",
        type=["xlsx"],
        help="Upload mechanic_skills_dataset.xlsx",
    )

    base_schedule_file = st.sidebar.file_uploader(
        "Base Aircraft Schedule",
        type=["xlsx"],
        help="Upload base_aircraft_schedule.xlsx",
    )

    cost_matrix_file = st.sidebar.file_uploader(
        "Cost Matrix",
        type=["xlsx"],
        help="Upload cost_matrix.xlsx",
    )

    avoidance_list_file = st.sidebar.file_uploader(
        "Avoidance List (Optional)",
        type=["xlsx"],
        help="Upload avoidance_list.xlsx (optional)",
    )

    if st.sidebar.button("🚀 Run Optimization", type="primary"):
        if not all([mechanic_skills_file, base_schedule_file, cost_matrix_file]):
            st.error(
                "❌ Please upload all required files: Mechanic Skills Dataset, "
                "Base Aircraft Schedule, and Cost Matrix"
            )
        else:
            progress_bar = st.progress(0)
            status_text = st.empty()

            try:
                # Load data
                status_text.text("📥 Loading data files...")
                progress_bar.progress(0.1)
                data = load_data(
                    mechanic_skills_file,
                    base_schedule_file,
                    cost_matrix_file,
                    avoidance_list_file,
                )
                progress_bar.progress(0.3)

                # Data summary
                with st.expander("📊 Data Summary", expanded=False):
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Mechanics", len(data["mechanics"]))
                    with col2:
                        st.metric("Bases", len(data["bases"]))
                    with col3:
                        st.metric("Periods", len(data["periods"]))
                    with col4:
                        st.metric("Shifts", len(data["shifts"]))

                # Create model
                status_text.text("⚙️ Creating optimization model...")
                progress_bar.progress(0.4)
                (
                    solver,
                    x,
                    mechanic_skills,
                    mechanic_inspector_skills,
                    inspector_req_columns,
                    avoidance_vars,
                ) = create_model(data)
                progress_bar.progress(0.7)

                # Model info
                with st.expander("🔧 Model Information", expanded=False):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Variables:** {solver.NumVariables()}")
                        st.write(f"**Constraints:** {solver.NumConstraints()}")
                    with col2:
                        st.write(f"**Solver:** {solver.SolverVersion()}")
                        if data["avoidance_dict"]:
                            unique_pairs = {
                                (min(k), max(k)) for k in data["avoidance_dict"].keys()
                            }
                            st.write(f"**Avoidance Pairs:** {len(unique_pairs)}")

                # Solve
                status, solve_time = solve_model(solver, progress_bar, status_text)

                if status in (pywraplp.Solver.OPTIMAL, pywraplp.Solver.FEASIBLE):
                    status_text.text("✅ Solution found! Extracting results...")

                    assignments, total_cost = extract_solution(
                        x,
                        data["mechanics"],
                        data["bases"],
                        data["periods"],
                        data["shifts"],
                        data["cost_dict"],
                    )

                    # Avoidance penalties
                    total_avoidance_penalty = 0
                    if data["avoidance_dict"] and avoidance_vars:
                        unique_pairs = {
                            (m1, m2)
                            for (m1, m2) in data["avoidance_dict"].keys()
                            if m1 < m2
                        }
                        for (m1, m2) in unique_pairs:
                            penalty = data["avoidance_dict"][(m1, m2)]
                            for b in data["bases"]:
                                for g in data["periods"]:
                                    for s in data["shifts"]:
                                        key = (m1, m2, b, g, s)
                                        if key in avoidance_vars and avoidance_vars[
                                            key
                                        ].solution_value() > 0.5:
                                            total_avoidance_penalty += penalty

                    objective_value = solver.Objective().Value()

                    st.success("✅ Optimization completed successfully!")

                    # Metrics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Total Assignments", len(assignments))
                    with col2:
                        st.metric("Movement Cost", f"${total_cost:.2f}")
                    with col3:
                        st.metric(
                            "Avoidance Penalty",
                            f"${total_avoidance_penalty:.2f}",
                        )
                    with col4:
                        st.metric("Total Objective", f"${objective_value:.2f}")

                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric(
                            "Unassigned Mechanics",
                            len(data["mechanics"]) - len(assignments),
                        )
                    with col2:
                        st.metric("Solve Time", f"{solve_time:.2f}s")

                    # Excel generation
                    status_text.text("📝 Generating Excel output...")
                    wb = generate_excel_output(
                        assignments,
                        data,
                        mechanic_skills,
                        mechanic_inspector_skills,
                        inspector_req_columns,
                        data["base_schedule_df"],
                    )

                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)

                    status_text.text("✅ Ready for download!")
                    progress_bar.empty()

                    st.download_button(
                        label="📥 Download Roster Output (Excel)",
                        data=output,
                        file_name=f"roster_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                    with st.expander("📋 View Assignments", expanded=False):
                        assignments_df = pd.DataFrame(assignments)
                        st.dataframe(assignments_df, use_container_width=True)

                elif status == pywraplp.Solver.INFEASIBLE:
                    st.error(
                        "❌ Problem is infeasible - no solution exists that satisfies all constraints"
                    )
                    status_text.text("❌ Infeasible problem")
                else:
                    st.warning(f"⚠️ Solver status: {status}")
                    status_text.text(f"⚠️ Status: {status}")

            except Exception as e:  # pragma: no cover - debug aid
                st.error(f"❌ Error: {str(e)}")
                status_text.text("❌ Error occurred")
                import traceback

                with st.expander("Error Details"):
                    st.code(traceback.format_exc())


if __name__ == "__main__":
    main()

